import json
import datetime
import logging
from decimal import Decimal
from conexionBD import obtener_conexion

def serializar_datos(obj):

    if isinstance(obj, (datetime.datetime, datetime.date)):

        return obj.isoformat()

    if isinstance(obj, Decimal):

        return float(obj)

    raise TypeError(f"Object of type {type(obj).__name__} is not JSON serializable")


def obtener_datos_para_jugar(pin, nickname):

    conexion = None

    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            sql_partida = """

                SELECT p.id_partida, p.id_cuestionario, c.titulo

                FROM Partida p

                JOIN Cuestionario c ON p.id_cuestionario = c.id_cuestionario

                WHERE p.pin_acceso = %s AND p.estado = 'en_curso'

            """
            cursor.execute(sql_partida, (pin,))
            partida = cursor.fetchone()
            if not partida:

                raise ValueError("La partida no está activa o el PIN es incorrecto.")
            id_partida = partida['id_partida']

            id_cuestionario = partida['id_cuestionario']
            sql_participante = """

                SELECT id_participante

                FROM Participante

                WHERE id_partida = %s AND nombre_usuario_partida = %s AND estado = 'Activo'

            """

            cursor.execute(sql_participante, (id_partida, nickname))

            participante = cursor.fetchone()



            if not participante:

                raise ValueError("No estás registrado en esta partida.")



            id_participante = participante['id_participante']



            sql_preguntas = """

                SELECT p.id_pregunta, p.texto_pregunta, p.url_media, p.tipo_pregunta,

                       p.tiempo_limite_segundos, p.puntos, p.opcion_rpta,

                       r.id_respuesta, r.texto_respuesta, r.es_correcta

                FROM Pregunta p

                LEFT JOIN Respuesta r ON p.id_pregunta = r.id_pregunta

                WHERE p.id_cuestionario = %s

                ORDER BY p.orden, r.id_respuesta

            """

            cursor.execute(sql_preguntas, (id_cuestionario,))

            preguntas_raw = cursor.fetchall()



            preguntas_dict = {}

            for row in preguntas_raw:

                id_pregunta = row['id_pregunta']

                if id_pregunta not in preguntas_dict:

                    preguntas_dict[id_pregunta] = {

                        'id_pregunta': id_pregunta,

                        'texto': row['texto_pregunta'],

                        'media': row['url_media'],

                        'tipo': row['tipo_pregunta'],

                        'tiempo': row['tiempo_limite_segundos'],

                        'puntos': row['puntos'],

                        'opcion_rpta': row['opcion_rpta'],

                        'respuestas': []

                    }

                if row.get('id_respuesta'):

                    preguntas_dict[id_pregunta]['respuestas'].append({

                        'id_respuesta': row['id_respuesta'],

                        'texto': row['texto_respuesta'],

                        'es_correcta': bool(row['es_correcta'])

                    })



            quiz_data = {

                'id_partida': id_partida,

                'id_participante': id_participante,

                'titulo': partida['titulo'],

                'preguntas': list(preguntas_dict.values())

            }



            return quiz_data



    except Exception as e:

        logging.error(f"Error en obtener_datos_para_jugar: {e}")

        raise e

    finally:

        if conexion:

            conexion.close()


def obtener_orden_primera_pregunta(id_cuestionario):

    conexion = None

    try:

        conexion = obtener_conexion()

        with conexion.cursor() as cursor:

            cursor.execute("SELECT MIN(orden) as primer_orden FROM Pregunta WHERE id_cuestionario = %s", (id_cuestionario,))

            resultado = cursor.fetchone()

            return resultado['primer_orden'] if resultado and resultado['primer_orden'] is not None else None

    finally:

        if conexion:

            conexion.close()


def actualizar_partida_para_inicio(id_partida, primer_orden):
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            # 🔥 KEY CHANGE: Added ", inicio_pregunta = NOW()"
            cursor.execute("""
                UPDATE Partida
                SET estado = 'en_curso', 
                    fase = 'pregunta', 
                    pregunta_actual_orden = %s, 
                    fecha_inicio = NOW(),
                    inicio_pregunta = NOW() 
                WHERE id_partida = %s
            """, (primer_orden, id_partida))
            conexion.commit()
    finally:
        if conexion:
            conexion.close()


def obtener_datos_partida_profesor(pin):
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            # --- 1. OBTENER DATOS DE LA PARTIDA ---
            sql_partida = """
                SELECT
                    p.id_partida, p.id_cuestionario, p.pregunta_actual_orden, p.estado, p.fase,
                    c.titulo AS quiz_titulo,
                    (SELECT COUNT(*) FROM Pregunta WHERE id_cuestionario = p.id_cuestionario) AS total_preguntas
                FROM Partida p
                JOIN Cuestionario c ON p.id_cuestionario = c.id_cuestionario
                WHERE p.pin_acceso = %s
            """
            cursor.execute(sql_partida, (pin,))
            partida = cursor.fetchone()

            if not partida:
                return {'estado_juego': 'error', 'message': 'Partida no encontrada.'}

            id_partida = partida['id_partida']
            id_cuestionario = partida['id_cuestionario']


            cursor.execute("SELECT COUNT(*) as num_grupos FROM Participante WHERE id_partida = %s AND id_grupo IS NOT NULL", (id_partida,))
            resultado_grupos = cursor.fetchone()
            hay_grupos = resultado_grupos['num_grupos'] > 0

            # --- 3. OBTENER MARCADOR (TOP 5) ---
            if hay_grupos:
                # MODO EQUIPO: Sumamos puntajes por grupo
                sql_marcador = """
                    SELECT CONCAT('Equipo ', id_grupo) AS nombre_usuario_partida,
                           SUM(puntuacion_total) AS puntuacion_total
                    FROM Participante
                    WHERE id_partida = %s AND estado = 'Activo' AND id_grupo IS NOT NULL
                    GROUP BY id_grupo
                    ORDER BY puntuacion_total DESC
                    LIMIT 5
                """
            else:
                # MODO INDIVIDUAL: Ranking normal
                sql_marcador = """
                    SELECT nombre_usuario_partida, COALESCE(puntuacion_total, 0) AS puntuacion_total
                    FROM Participante
                    WHERE id_partida = %s AND estado = 'Activo'
                    ORDER BY puntuacion_total DESC
                    LIMIT 5
                """

            cursor.execute(sql_marcador, (id_partida,))
            marcador = cursor.fetchall()

            # --- 4. SI EL JUEGO ESTÁ FINALIZADO ---
            if partida['estado'] == 'finalizada':
                 cursor.execute("SELECT descripcion FROM Recompensa WHERE id_cuestionario = %s ORDER BY id_recompensa ASC", (id_cuestionario,))
                 recompensas_raw = cursor.fetchall()
                 recompensas = [r['descripcion'] for r in recompensas_raw if r.get('descripcion')]

                 return {
                    'estado_juego': 'finalizada',
                    'marcador': marcador,
                    'recompensas': recompensas
                 }

            # --- 5. SI EL JUEGO ESTÁ EN CURSO ---

            pregunta_actual_orden_num = partida['pregunta_actual_orden']

            # Obtenemos la lista de estudiantes para ver quién ha respondido
            cursor.execute("""
                SELECT COUNT(*) AS total, id_participante, nombre_usuario_partida
                FROM Participante WHERE id_partida = %s AND estado = 'Activo'
                GROUP BY id_participante
            """, (id_partida,))
            participantes = cursor.fetchall()
            total_participantes = len(participantes)

            # A) Si no hay pregunta activa (Pantalla de inicio o transición)
            if not pregunta_actual_orden_num:
                return {
                    'estado_juego': 'en_curso',
                    'fase': partida.get('fase', 'pregunta'),
                    'pregunta_actual': None,
                    'respuestas_info': {'recibidas': 0, 'total': total_participantes, 'estadisticas': {}},
                    'marcador': marcador,
                    'estudiantes': [{'nombre_usuario_partida': p['nombre_usuario_partida'], 'ha_respondido': False} for p in participantes]
                }

            # B) Hay una pregunta activa, obtenemos sus datos
            sql_pregunta = """
                SELECT id_pregunta, texto_pregunta, tiempo_limite_segundos, tipo_pregunta, orden
                FROM Pregunta
                WHERE id_cuestionario = %s AND orden = %s
            """
            cursor.execute(sql_pregunta, (id_cuestionario, pregunta_actual_orden_num))
            pregunta = cursor.fetchone()

            if not pregunta:
                return {'estado_juego': 'error', 'message': f'Error: No se encontró la pregunta {pregunta_actual_orden_num}'}

            id_pregunta_actual = pregunta['id_pregunta']

            # C) Obtener todas las respuestas posibles (para el gráfico)
            cursor.execute("""
                SELECT id_respuesta, texto_respuesta
                FROM Respuesta
                WHERE id_pregunta = %s
                ORDER BY id_respuesta
            """, (id_pregunta_actual,))
            respuestas_completas = cursor.fetchall()

            # D) Contar respuestas recibidas en tiempo real
            sql_respuestas_recibidas = """
                SELECT r.texto_respuesta, rp.id_participante
                FROM RespuestaParticipante rp
                JOIN Respuesta r ON rp.id_respuesta_seleccionada = r.id_respuesta
                WHERE rp.id_pregunta = %s
            """
            cursor.execute(sql_respuestas_recibidas, (id_pregunta_actual,))
            respuestas_raw = cursor.fetchall()

            participantes_que_respondieron = {r['id_participante'] for r in respuestas_raw}
            respuestas_recibidas_conteo = len(participantes_que_respondieron)

            estadisticas_conteo = {}
            for r in respuestas_raw:
                texto = r['texto_respuesta']
                estadisticas_conteo[texto] = estadisticas_conteo.get(texto, 0) + 1

            estudiantes_status = []
            for p in participantes:
                estudiantes_status.append({
                    'nombre_usuario_partida': p['nombre_usuario_partida'],
                    'ha_respondido': p['id_participante'] in participantes_que_respondieron
                })

            return {
                'estado_juego': 'en_curso',
                'fase': partida.get('fase', 'pregunta'),
                'codigo_partida': pin,
                'quiz_titulo': partida['quiz_titulo'],
                'pregunta_actual': {
                    'id_pregunta': pregunta['id_pregunta'],
                    'numero': pregunta_actual_orden_num,
                    'total_preguntas': partida['total_preguntas'],
                    'categoria': pregunta.get('tipo_pregunta', 'General'),

                    # --- ESTA ES LA CORRECCIÓN CLAVE ---
                    'texto': pregunta['texto_pregunta'],
                    # -----------------------------------

                    'tiempo_limite': pregunta['tiempo_limite_segundos'],
                    'respuestas': respuestas_completas
                },
                'respuestas_info': {
                    'recibidas': respuestas_recibidas_conteo,
                    'total': total_participantes,
                    'estadisticas': estadisticas_conteo
                },
                'marcador': marcador,
                'estudiantes': estudiantes_status
            }

    except Exception as e:
        import logging
        logging.error(f"Error en obtener_datos_partida_profesor: {e}")
        return {'estado_juego': 'error', 'message': str(e)}

    finally:
        if conexion:
            conexion.close()


def avanzar_siguiente_pregunta(pin):
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            cursor.execute("SELECT id_partida, id_cuestionario, pregunta_actual_orden FROM Partida WHERE pin_acceso = %s", (pin,))
            partida = cursor.fetchone()

            if not partida:
                return {'success': False, 'message': 'Partida no encontrada'}

            orden_actual = partida['pregunta_actual_orden'] or 0

            cursor.execute("SELECT MIN(orden) AS siguiente_orden FROM Pregunta WHERE id_cuestionario = %s AND orden > %s", (partida['id_cuestionario'], orden_actual))
            siguiente = cursor.fetchone()

            if siguiente and siguiente['siguiente_orden'] is not None:
                nuevo_orden = siguiente['siguiente_orden']
                
                # 🔥 KEY CHANGE: Updates "inicio_pregunta = NOW()"
                cursor.execute("""
                    UPDATE Partida
                    SET pregunta_actual_orden = %s, 
                        fase = 'pregunta', 
                        inicio_pregunta = NOW() 
                    WHERE id_partida = %s
                """, (nuevo_orden, partida['id_partida']))

            else:
                # Game Finished logic
                cursor.execute("""
                    UPDATE Partida
                    SET estado = 'finalizada', fase = 'finalizada', pregunta_actual_orden = NULL, fecha_fin = NOW()
                    WHERE id_partida = %s
                """, (partida['id_partida'],))

            conexion.commit()
            return {'success': True}

    except Exception as e:
        if conexion: conexion.rollback()
        logging.error(f"Error en avanzar_siguiente_pregunta: {e}")
        return {'success': False, 'message': str(e)}
    finally:
        if conexion:
            conexion.close()

def obtener_estado_actual_partida(pin, nickname):
    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            # 🔥 KEY CHANGE: Select "inicio_pregunta" instead of generic date
            cursor.execute("""
                SELECT id_partida, id_cuestionario, estado, fase, pregunta_actual_orden, inicio_pregunta
                FROM Partida WHERE pin_acceso = %s
            """, (pin,))
            partida = cursor.fetchone()

            if not partida: return {'estado_juego': 'finalizada'}

            id_partida = partida['id_partida']
            id_cuestionario = partida['id_cuestionario']

            # --- CALCULATE EXACT ELAPSED TIME ---
            segundos_transcurridos = 0
            
            # Only calculate if game is active, in question phase, and start time exists
            if partida['estado'] == 'en_curso' and partida['fase'] == 'pregunta' and partida['inicio_pregunta']:
                ahora = datetime.datetime.now()
                inicio_dt = partida['inicio_pregunta']
                
                # Robust date parsing (handles string vs datetime object)
                if isinstance(inicio_dt, str):
                    try:
                        inicio_dt = datetime.datetime.fromisoformat(str(inicio_dt))
                    except:
                        try:
                            # Fallback to SQL standard format
                            inicio_dt = datetime.datetime.strptime(str(inicio_dt), '%Y-%m-%d %H:%M:%S')
                        except:
                            inicio_dt = ahora 
                
                try:
                    delta = ahora - inicio_dt
                    segundos_transcurridos = delta.total_seconds()
                except:
                    segundos_transcurridos = 0
            
            segundos_transcurridos = max(0, int(segundos_transcurridos))

            # --- FETCH STUDENT DATA ---
            cursor.execute("""
                SELECT id_participante, id_grupo, COALESCE(puntuacion_total, 0) as score_individual
                FROM Participante
                WHERE id_partida = %s AND nombre_usuario_partida = %s AND estado = 'Activo'
            """, (id_partida, nickname))
            participante = cursor.fetchone()

            if not participante:
                if partida['estado'] == 'finalizada' or partida['fase'] == 'finalizada':
                    id_participante = -1
                    mi_score_individual = 0
                else:
                    return {'estado_juego': 'expulsado', 'message': 'Ya no estás en esta partida.'}
            else:
                id_participante = participante['id_participante']
                mi_score_individual = participante['score_individual']

            # --- GAME FINISHED LOGIC ---
            if partida['estado'] == 'finalizada' or partida['fase'] == 'finalizada':
                cursor.execute("SELECT descripcion FROM Recompensa WHERE id_cuestionario = %s ORDER BY id_recompensa ASC", (id_cuestionario,))
                recompensas = [r['descripcion'] for r in cursor.fetchall() if r.get('descripcion')]

                cursor.execute("""
                    SELECT nombre_usuario_partida, COALESCE(puntuacion_total, 0) AS puntuacion_total
                    FROM Participante WHERE id_partida = %s AND estado = 'Activo'
                    ORDER BY puntuacion_total DESC LIMIT 5
                """, (id_partida,))
                marcador = cursor.fetchall()

                cursor.execute("""
                    SELECT COUNT(*) + 1 AS puesto FROM Participante 
                    WHERE id_partida = %s AND puntuacion_total > %s
                """, (id_partida, mi_score_individual))
                puesto_data = cursor.fetchone()
                puesto = puesto_data['puesto'] if puesto_data else 0

                return {
                    'estado_juego': 'finalizada',
                    'fase': 'finalizada',
                    'id_participante': id_participante,
                    'puesto': puesto,
                    'puntuacion_final': mi_score_individual,
                    'puntuacion_individual': mi_score_individual,
                    'marcador': marcador,
                    'recompensas': recompensas
                }

            # --- GAME IN PROGRESS LOGIC ---
            pregunta_actual_orden_num = partida['pregunta_actual_orden']
            
            sql_pregunta = """
                SELECT p.id_pregunta, p.texto_pregunta AS texto, p.url_media AS media,
                    p.tiempo_limite_segundos AS tiempo, p.opcion_rpta, p.orden AS numero,
                    (SELECT COUNT(*) FROM Pregunta WHERE id_cuestionario = p.id_cuestionario) AS total_preguntas
                FROM Pregunta p WHERE p.id_cuestionario = %s AND p.orden = %s
            """
            cursor.execute(sql_pregunta, (id_cuestionario, pregunta_actual_orden_num))
            pregunta_actual = cursor.fetchone()

            respuesta_enviada = None
            if pregunta_actual:
                id_pregunta = pregunta_actual['id_pregunta']
                cursor.execute("SELECT id_respuesta, texto_respuesta FROM Respuesta WHERE id_pregunta = %s", (id_pregunta,))
                pregunta_actual['respuestas'] = cursor.fetchall()

                cursor.execute("""
                    SELECT r.es_correcta, rp.puntuacion_obtenida
                    FROM RespuestaParticipante rp
                    JOIN Respuesta r ON rp.id_respuesta_seleccionada = r.id_respuesta
                    WHERE rp.id_participante = %s AND rp.id_pregunta = %s
                """, (id_participante, id_pregunta))
                respuesta_enviada = cursor.fetchone()

            # Return JSON with calculated elapsed time
            return json.loads(json.dumps({
                'estado_juego': 'en_curso',
                'fase': partida.get('fase', 'pregunta'),
                # ✅ Sending correct elapsed time calculated from inicio_pregunta
                'segundos_transcurridos': segundos_transcurridos, 
                'pregunta_actual': pregunta_actual,
                'resultado_respuesta': respuesta_enviada,
                'id_participante': id_participante,
                'puntuacion_individual': mi_score_individual
            }, default=serializar_datos))

    except Exception as e:
        logging.error(f"Error en obtener_estado_actual_partida: {e}", exc_info=True)
        return {'estado_juego': 'finalizada'}
    finally:
        if conexion: conexion.close()


def guardar_respuesta_alumno(id_participante, id_pregunta, id_respuesta_seleccionada, tiempo_respuesta_segundos):

    conexion = None

    puntos_calculados = 0

    es_correcta = False

    try:

        conexion = obtener_conexion()

        with conexion.cursor() as cursor:

            cursor.execute("""

                SELECT 1 FROM RespuestaParticipante

                WHERE id_participante = %s AND id_pregunta = %s

            """, (id_participante, id_pregunta))

            if cursor.fetchone():

                return {"success": False, "message": "Ya respondiste esta pregunta."}



            cursor.execute("""

                SELECT pr.puntos AS puntos_pregunta, pr.tiempo_limite_segundos, r.es_correcta

                FROM Respuesta r

                JOIN Pregunta pr ON r.id_pregunta = pr.id_pregunta

                WHERE r.id_respuesta = %s AND r.id_pregunta = %s

            """, (id_respuesta_seleccionada, id_pregunta))

            info = cursor.fetchone()



            puntos_base_pregunta = 'Estándar'

            tiempo_limite = 30



            if info:

                es_correcta = bool(info['es_correcta'])

                puntos_base_pregunta = info['puntos_pregunta']

                tiempo_limite = info.get('tiempo_limite_segundos', 30)



            if es_correcta:

                PUNTOS_BASE = 1000

                try:

                    tiempo_respuesta_float = float(tiempo_respuesta_segundos)

                    tiempo_limite_float = float(tiempo_limite)

                    factor_tiempo = max(0.5, (1 - (tiempo_respuesta_float / tiempo_limite_float) / 2))

                    puntos_calculados = round(PUNTOS_BASE * factor_tiempo)

                except:

                    puntos_calculados = PUNTOS_BASE // 2



                if puntos_base_pregunta == 'Puntos dobles':

                    puntos_calculados *= 2

                elif puntos_base_pregunta == 'Sin puntos':

                    puntos_calculados = 0

            else:

                puntos_calculados = 0



            logging.debug(f"Puntos calculados para participante {id_participante}: {puntos_calculados}")

            cursor.execute("""

                INSERT INTO RespuestaParticipante

                (id_participante, id_pregunta, id_respuesta_seleccionada, tiempo_respuesta_segundos, puntuacion_obtenida)

                VALUES (%s, %s, %s, %s, %s)

            """, (id_participante, id_pregunta, id_respuesta_seleccionada, tiempo_respuesta_segundos, puntos_calculados))



            if puntos_calculados > 0:

                cursor.execute("""

                    UPDATE Participante

                    SET puntuacion_total = COALESCE(puntuacion_total, 0) + %s, fecha_ultima_respuesta = NOW()

                    WHERE id_participante = %s

                """, (puntos_calculados, id_participante))



            conexion.commit()



            return {

                "success": True,

                "es_correcta": es_correcta,

                "puntos_obtenidos": puntos_calculados

            }



    except Exception as e:

        logging.error(f"Error CRÍTICO en guardar_respuesta_alumno: {e}")

        if conexion:

            conexion.rollback()

        return {

            "success": False,

            "es_correcta": False,

            "puntos_obtenidos": 0,

            "error": str(e)

        }



    finally:

        if conexion:

            conexion.close()


def guardar_puntuacion_final(id_participante, puntuacion_total):

    conexion = None

    try:

        conexion = obtener_conexion()

        with conexion.cursor() as cursor:

            cursor.execute("""

                UPDATE Participante SET puntuacion_total = %s WHERE id_participante = %s

            """, (puntuacion_total, id_participante))

            conexion.commit()

    except Exception as e:

        logging.error(f"Error en guardar_puntuacion_final: {e}")

        if conexion: conexion.rollback()

    finally:

        if conexion:

            conexion.close()


def forzar_fase_resultado(pin):

    conexion = None

    try:

        conexion = obtener_conexion()

        with conexion.cursor() as cursor:

            cursor.execute("""

                UPDATE Partida

                SET fase = 'resultado'

                WHERE pin_acceso = %s AND estado = 'en_curso'

            """, (pin,))

            conexion.commit()



        logging.info(f"✅ Fase de la partida con PIN {pin} forzada a 'resultado'")

        return {'success': True}



    except Exception as e:

        logging.error(f"❌ Error en forzar_fase_resultado({pin}): {e}")

        if conexion:

            conexion.rollback()

        return {'success': False, 'message': str(e)}



    finally:

        if conexion:

            conexion.close()


def notificar_cambio_partida(pin):

    try:

        from app import partida_events

        if pin in partida_events:

            partida_events[pin].set()

            logging.info(f"🔔 Notificación enviada a los alumnos del PIN {pin}")

        else:

            logging.warning(f"⚠️ No hay evento activo para el PIN {pin}")



    except Exception as e:

        logging.error(f"❌ Error en notificar_cambio_partida({pin}): {e}")


def obtener_pin_por_participante(id_participante):

    conexion = None

    try:

        conexion = obtener_conexion()

        with conexion.cursor() as cursor:

            cursor.execute("""

                SELECT pa.pin_acceso

                FROM Partida pa

                JOIN Participante p ON p.id_partida = pa.id_partida

                WHERE p.id_participante = %s

            """, (id_participante,))

            fila = cursor.fetchone()

            if fila:

                return fila['pin_acceso']

            return None

    except Exception as e:

        logging.error(f"Error en obtener_pin_por_participante: {e}")

        return None

    finally:

        if conexion:

            conexion.close()

def exportar_resultados_excel(pin):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    conexion = None
    try:
        conexion = obtener_conexion()
        with conexion.cursor() as cursor:
            # 1) Obtener datos básicos
            cursor.execute("SELECT id_partida, id_cuestionario FROM Partida WHERE pin_acceso = %s", (pin,))
            partida = cursor.fetchone()
            if not partida: raise ValueError("Partida no encontrada.")

            id_partida = partida['id_partida']
            id_cuestionario = partida['id_cuestionario']

            # 2) DETECTAR SI HAY GRUPOS (Esta es la clave)
            cursor.execute("SELECT COUNT(*) as num_grupos FROM Participante WHERE id_partida = %s AND id_grupo IS NOT NULL", (id_partida,))
            hay_grupos = cursor.fetchone()['num_grupos'] > 0

            # 3) Obtener Preguntas (Columnas)
            cursor.execute("SELECT id_pregunta, texto_pregunta, orden FROM Pregunta WHERE id_cuestionario = %s ORDER BY orden ASC", (id_cuestionario,))
            preguntas = cursor.fetchall()

            # 4) Obtener Totales por Grupo (Solo si es grupal)
            mapa_totales_grupo = {}
            if hay_grupos:
                cursor.execute("""
                    SELECT id_grupo, SUM(puntuacion_total) as total_grupo
                    FROM Participante WHERE id_partida = %s AND id_grupo IS NOT NULL GROUP BY id_grupo
                """, (id_partida,))
                totales = cursor.fetchall()
                for t in totales: mapa_totales_grupo[t['id_grupo']] = t['total_grupo']

            # 5) Obtener Participantes (Orden dinámico)
            orden_sql = "id_grupo, puntuacion_total DESC" if hay_grupos else "puntuacion_total DESC"
            cursor.execute(f"""
                SELECT id_participante, nombre_usuario_partida, COALESCE(puntuacion_total, 0) AS puntuacion_total, id_grupo
                FROM Participante
                WHERE id_partida = %s
                ORDER BY {orden_sql}
            """, (id_partida,))
            participantes = cursor.fetchall()

            # 6) Obtener Respuestas
            cursor.execute("""
                SELECT rp.id_participante, rp.id_pregunta, r.es_correcta
                FROM RespuestaParticipante rp
                JOIN Respuesta r ON r.id_respuesta = rp.id_respuesta_seleccionada
                WHERE rp.id_participante IN (SELECT id_participante FROM Participante WHERE id_partida = %s)
            """, (id_partida,))
            respuestas = cursor.fetchall()

            mapa_resp = {}
            for row in respuestas:
                mapa_resp[(row['id_participante'], row['id_pregunta'])] = bool(row['es_correcta'])

            # ==========================================
            # 🎨 CONSTRUCCIÓN DEL EXCEL
            # ==========================================
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados"

            # --- ESTILOS ---
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul
            header_font = Font(bold=True, color="FFFFFF", size=12)

            # --- ENCABEZADOS DE COLUMNA ---
            headers = ["Nombre Participante", "Puntaje Individual"]
            for preg in preguntas:
                headers.append(f"P{preg['orden']}")

            ws.append(headers)

            # Estilizar cabecera
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                width = 25 if col_idx == 1 else (18 if col_idx == 2 else 8)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            current_row = 2

            # =========================================================
            # 🔴 LÓGICA DIVIDIDA: INDIVIDUAL vs GRUPAL
            # =========================================================

            if not hay_grupos:
                # 👉 OPCIÓN A: MODO INDIVIDUAL CLÁSICO (Lista plana)
                for p in participantes:
                    row_data = [p['nombre_usuario_partida'], p['puntuacion_total']]
                    # Respuestas
                    for preg in preguntas:
                        key = (p['id_participante'], preg['id_pregunta'])
                        val = "Correcta" if mapa_resp.get(key) else ("Incorrecta" if key in mapa_resp else "-")
                        row_data.append(val)

                    # Escribir fila normal sin estilos raros
                    for col_idx, val in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        # Alinear centro las respuestas y puntajes
                        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")

                    current_row += 1

            else:
                # 👉 OPCIÓN B: MODO GRUPAL (Con encabezados bonitos)

                # Estilos extra solo para grupos
                group_header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                group_header_font = Font(bold=True, color="000000", size=11)

                # Organizar datos
                grupos_data = {}
                sin_grupo_data = []
                for p in participantes:
                    gid = p['id_grupo']
                    if gid:
                        if gid not in grupos_data: grupos_data[gid] = []
                        grupos_data[gid].append(p)
                    else:
                        sin_grupo_data.append(p)

                # Escribir Grupos
                for gid in sorted(grupos_data.keys()):
                    miembros = grupos_data[gid]
                    total_equipo = mapa_totales_grupo.get(gid, 0)

                    # Título del Grupo (Celda combinada amarilla)
                    titulo_grupo = f"🏆 EQUIPO {gid}  ---  RESULTADO TOTAL: {total_equipo} PUNTOS"
                    ws.cell(row=current_row, column=1, value=titulo_grupo)
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))

                    for col in range(1, len(headers) + 1):
                        c = ws.cell(row=current_row, column=col)
                        c.fill = group_header_fill
                        c.font = group_header_font
                        c.border = Border(bottom=Side(style='thin'))

                    current_row += 1

                    # Miembros del grupo
                    for p in miembros:
                        row_data = [p['nombre_usuario_partida'], p['puntuacion_total']]
                        for preg in preguntas:
                            key = (p['id_participante'], preg['id_pregunta'])
                            val = "Correcta" if mapa_resp.get(key) else ("Incorrecta" if key in mapa_resp else "-")
                            row_data.append(val)

                        for col_idx, val in enumerate(row_data, 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=val)
                            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")
                        current_row += 1

                    current_row += 1 # Espacio entre grupos

                # Escribir "Sin Equipo" (Solo si hay grupos Y gente suelta)
                if sin_grupo_data:
                    ws.cell(row=current_row, column=1, value="SIN EQUIPO").font = Font(bold=True)
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
                    current_row += 1

                    for p in sin_grupo_data:
                        row_data = [p['nombre_usuario_partida'], p['puntuacion_total']]
                        for preg in preguntas:
                            key = (p['id_participante'], preg['id_pregunta'])
                            val = "Correcta" if mapa_resp.get(key) else ("Incorrecta" if key in mapa_resp else "-")
                            row_data.append(val)

                        for col_idx, val in enumerate(row_data, 1):
                            ws.cell(row=current_row, column=col_idx, value=val)
                        current_row += 1

            # =========================================================

            ws.freeze_panes = "A2"
            output = io.BytesIO()
            wb.save(output)

            tipo = "equipos" if hay_grupos else "resultados"
            filename = f"{tipo}_{pin}.xlsx"
            return output, filename

    except Exception as e:
        import logging
        logging.error(f"Error en exportar_resultados_excel({pin}): {e}", exc_info=True)
        raise
    finally:
        if conexion: conexion.close()

# En controladorPartida.py

def obtener_estadisticas_pregunta(id_partida, id_pregunta):
    conexion = obtener_conexion()
    estadisticas = {}
    try:
        with conexion.cursor() as cursor:
            # 1. Obtener todas las respuestas posibles de esa pregunta
            cursor.execute("SELECT id_respuesta, texto_respuesta FROM Respuesta WHERE id_pregunta = %s", (id_pregunta,))
            respuestas_posibles = cursor.fetchall()

            # 2. Inicializar contador en 0 para todas
            for r in respuestas_posibles:
                estadisticas[r['texto_respuesta']] = 0

            # 3. Contar respuestas reales de los participantes
            query = """
                SELECT r.texto_respuesta, COUNT(rp.id_respuesta_participante) as total
                FROM RespuestaParticipante rp
                JOIN Respuesta r ON rp.id_respuesta_seleccionada = r.id_respuesta
                WHERE rp.id_participante IN (SELECT id_participante FROM Participante WHERE id_partida = %s)
                AND rp.id_pregunta = %s
                GROUP BY r.texto_respuesta
            """
            cursor.execute(query, (id_partida, id_pregunta))
            resultados = cursor.fetchall()

            # 4. Actualizar el diccionario con los conteos reales
            for res in resultados:
                estadisticas[res['texto_respuesta']] = res['total']

    except Exception as e:
        print(f"Error obteniendo estadísticas: {e}")
    finally:
        conexion.close()

    return estadisticas